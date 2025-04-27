from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl.cell.cell import MergedCell

import os

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(BASE_DIR, 'data.xlsx')

green_fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")  # أخضر غامق
orange_fill = PatternFill(start_color="EF6C00", end_color="EF6C00", fill_type="solid") # برتقالي غامق
yellow_fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid") # أصفر
red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")    # أحمر غامق
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # أبيض
@app.route('/')
def index():
    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheets_data = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value if cell.value else '' for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            formatted_row = [{'value': str(cell) if cell is not None else ''} for cell in row]
            rows.append(formatted_row)
        sheets_data.append({'name': sheet, 'headers': headers, 'rows': rows})

    return render_template('index.html', sheets_data=sheets_data)

def apply_styles(sheet, row_idx):
    title = sheet.title.strip()
    headers = [cell.value for cell in sheet[1]]

    if title == "متابعة المشاريع وزارة التخطيط":
        for cell in sheet[row_idx]:
            value = str(cell.value).strip() if cell.value else ""
            if value == "تم الإعلان":
                cell.fill = green_fill
                if cell.column + 1 <= sheet.max_column:
                    sheet.cell(row=row_idx, column=cell.column + 1).fill = white_fill
            elif value == "تم":
                cell.fill = green_fill
                next_cell = sheet.cell(row=row_idx, column=cell.column + 1)
                next_value = str(next_cell.value).strip() if next_cell.value else ""
                if not next_value:
                    next_cell.fill = orange_fill
                else:
                    try:
                        date = datetime.strptime(next_value.split()[0], "%Y-%m-%d")
                        today = datetime.now()
                        if date.date() < today.date():
                            next_cell.fill = green_fill
                        else:
                            next_cell.fill = yellow_fill
                    except:
                        pass

    elif title in ["متابعة مشاريع قيد التنفيذ"]:
        if "نسبة الانحراف %" in headers:
            idx = headers.index("نسبة الانحراف %") + 1
            cell = sheet.cell(row=row_idx, column=idx)
            try:
                num = float(cell.value)
                if num < 0:
                    cell.fill = red_fill
                elif num > 0:
                    cell.fill = green_fill
                else:
                    cell.fill = white_fill
            except:
                pass

    elif title in ["متابعة التوقفات", "متابعة المدد الاضافية", "تحديث وامر الغيار"]:
        if "تاريخ الإنجاز المتوقع" in headers:
            idx = headers.index("تاريخ الإنجاز المتوقع") + 1
            cell = sheet.cell(row=row_idx, column=idx)
            prev_cell = sheet.cell(row=row_idx - 1, column=idx) if row_idx > 2 else None
            try:
                today = datetime.today().date()
                current_date = datetime.strptime(str(cell.value).split()[0], "%Y-%m-%d").date()
                if prev_cell and prev_cell.value:
                    prev_date = datetime.strptime(str(prev_cell.value).split()[0], "%Y-%m-%d").date()
                    if current_date > prev_date:
                        cell.fill = orange_fill
                if current_date > today:
                    cell.fill = red_fill
            except:
                pass
@app.route('/save', methods=['POST'])
def save():
    try:
        data = request.get_json()  # استخدام get_json أو request.json ماكو فرق كبير هنا

        wb = load_workbook(EXCEL_FILE_PATH)

        for sheet_name, rows in data.items():
            if sheet_name not in wb.sheetnames:
                continue  # إذا اسم الشيت مو موجود، تجاهله بدل ما يصير خطأ

            sheet = wb[sheet_name]
            for row_index, row in enumerate(rows, start=2):
                for col_index, value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_index, column=col_index)
                    if not isinstance(cell, MergedCell):  # ✅ تأكد مو خلية مدموجة
                        if value in ("None", None):
                            value = ""
                        cell.value = value
                apply_styles(sheet, row_index)
                # لو عندك تنسيقات أو تلوين للخلايا
                apply_styles(sheet, row_index)

        wb.save(EXCEL_FILE_PATH)
        return jsonify({'status': 'success'})

    except Exception as e:
        print(f"Error while saving: {e}")  # تطبع الخطأ باللوج للتسهيل
        return jsonify({'status': 'error', 'message': str(e)}), 500
@app.route('/export')
def export():
    return send_file(EXCEL_FILE_PATH, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
