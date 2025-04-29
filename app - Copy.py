from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl.cell.cell import MergedCell
import os

app = Flask(__name__)

# ğŸ”µ Ø§Ù„Ù…Ø³Ø§Ø± Ø§Ù„Ø«Ø§Ø¨Øª Ù„Ù…Ù„Ù Ø§Ù„Ø¥ÙƒØ³Ù„
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(BASE_DIR, 'data.xlsx')

# ğŸ”µ ØªØ¹Ø±ÙŠÙ Ø£Ù„ÙˆØ§Ù† Ø§Ù„ØªÙ„ÙˆÙŠÙ†
green_fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
orange_fill = PatternFill(start_color="EF6C00", end_color="EF6C00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")
red_fill = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

@app.route('/')
def index():
    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheets_data = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value if cell.value else '' for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            formatted_row = [{'value': str(cell) if cell is not None else ''} for cell in row]
            rows.append(formatted_row)
        sheets_data.append({'name': sheet, 'headers': headers, 'rows': rows})

    return render_template('index.html', sheets_data=sheets_data)

@app.route('/save', methods=['POST'])
def save():
    try:
        data = request.get_json()
        wb = load_workbook(EXCEL_FILE_PATH)

        for sheet_name, rows in data.items():
            if sheet_name not in wb.sheetnames:
                continue

            sheet = wb[sheet_name]
            for row_index, row in enumerate(rows[1:], start=2):  # ØªØ¬Ø§Ù‡Ù„ Ø§Ù„Ø¹Ù†Ø§ÙˆÙŠÙ†
                for col_index, value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_index, column=col_index)
                    if not isinstance(cell, MergedCell):
                        cell.value = "" if value in ("None", None) else value
                apply_styles(sheet, row_index)

        wb.save(EXCEL_FILE_PATH)
        return jsonify({'status': 'success'})

    except Exception as e:
        print(f"Error while saving: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/export')
def export():
    return send_file(EXCEL_FILE_PATH, as_attachment=True)

@app.route('/report-fields')
def report_fields():
    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheets_data = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value if cell.value else '' for cell in ws[1]]
        sheets_data.append({'name': sheet, 'headers': headers})

    return render_template('report_fields.html', sheets_data=sheets_data)

def apply_styles(sheet, row_idx):
    title = sheet.title.strip()
    headers = [cell.value for cell in sheet[1]]

    # ğŸ”µ Ø´ÙŠØª Ù…ØªØ§Ø¨Ø¹Ø© Ø§Ù„Ù…Ø´Ø§Ø±ÙŠØ¹ ÙˆØ²Ø§Ø±Ø© Ø§Ù„ØªØ®Ø·ÙŠØ·
    if title == "Ø´ÙŠØª Ù…ØªØ§Ø¨Ø¹Ø© Ø§Ù„Ù…Ø´Ø§Ø±ÙŠØ¹ ÙˆØ²Ø§Ø±Ø© Ø§Ù„ØªØ®Ø·ÙŠØ·":
        for col_idx, cell in enumerate(sheet[row_idx], start=1):
            value = str(cell.value).strip() if cell.value else ''
            next_cell = sheet.cell(row=row_idx, column=col_idx + 1)

            if value == "ØªÙ… Ø§Ù„Ø¥Ø¹Ù„Ø§Ù†":
                cell.fill = green_fill
                if next_cell:
                    next_cell.fill = white_fill

            elif value == "ØªÙ…":
                cell.fill = green_fill
                if next_cell:
                    next_value = str(next_cell.value).strip() if next_cell.value else ''
                    if next_value == "":
                        next_cell.fill = orange_fill
                    elif next_value.count("-") == 2:
                        try:
                            cell_date = datetime.strptime(next_value, "%Y-%m-%d").date()
                            today = datetime.today().date()
                            if cell_date < today:
                                next_cell.fill = green_fill
                            elif cell_date > today:
                                next_cell.fill = yellow_fill
                        except:
                            next_cell.fill = white_fill
                    else:
                        next_cell.fill = white_fill
            else:
                cell.fill = white_fill

    # ğŸ”µ Ø´ÙŠØª Ù…ØªØ§Ø¨Ø¹Ø© Ù…Ø´Ø§Ø±ÙŠØ¹ Ù‚ÙŠØ¯ Ø§Ù„ØªÙ†ÙÙŠØ°
    elif title == "Ù…ØªØ§Ø¨Ø¹Ø© Ù…Ø´Ø§Ø±ÙŠØ¹ Ù‚ÙŠØ¯ Ø§Ù„ØªÙ†ÙÙŠØ°":
        if "Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ù†Ø­Ø±Ø§Ù %" in headers:
            idx = headers.index("Ù†Ø³Ø¨Ø© Ø§Ù„Ø§Ù†Ø­Ø±Ø§Ù %") + 1
            cell = sheet.cell(row=row_idx, column=idx)
            try:
                num = float(cell.value)
                if num < 0:
                    cell.fill = red_fill
                elif num > 0:
                    cell.fill = green_fill
                else:
                    cell.fill = white_fill
            except:
                cell.fill = white_fill

    # ğŸ”µ Ø´ÙŠØª Ù…ØªØ§Ø¨Ø¹Ø© Ø§Ù„ØªÙˆÙ‚ÙØ§Øª ÙˆØ§Ù„Ù…Ø¯Ø¯ Ø§Ù„Ø¥Ø¶Ø§ÙÙŠØ© ÙˆØ£Ù…Ø± Ø§Ù„ØºÙŠØ§Ø±
    elif title in ["Ù…ØªØ§Ø¨Ø¹Ø© Ø§Ù„ØªÙˆÙ‚ÙØ§Øª", "Ù…ØªØ§Ø¨Ø¹Ø© Ø§Ù„Ù…Ø¯Ø¯ Ø§Ù„Ø§Ø¶Ø§ÙÙŠØ©", "ØªØ­Ø¯ÙŠØ« ÙˆØ§Ù…Ø± Ø§Ù„ØºÙŠØ§Ø±"]:
        if "ØªØ§Ø±ÙŠØ® Ø§Ù„Ø¥Ù†Ø¬Ø§Ø² Ø§Ù„Ù…ØªÙˆÙ‚Ø¹" in headers:
            idx = headers.index("ØªØ§Ø±ÙŠØ® Ø§Ù„Ø¥Ù†Ø¬Ø§Ø² Ø§Ù„Ù…ØªÙˆÙ‚Ø¹") + 1
            cell = sheet.cell(row=row_idx, column=idx)
            prev_cell = sheet.cell(row=row_idx - 1, column=idx) if row_idx > 2 else None
            try:
                today = datetime.today().date()
                current_date = datetime.strptime(str(cell.value).split()[0], "%Y-%m-%d").date()

                if prev_cell and prev_cell.value:
                    prev_date = datetime.strptime(str(prev_cell.value).split()[0], "%Y-%m-%d").date()
                    if current_date > prev_date:
                        cell.fill = orange_fill  # ØªØ£Ø®ÙŠØ± Ø¹Ù† Ø§Ù„Ø³Ø·Ø± Ø§Ù„Ø³Ø§Ø¨Ù‚

                if current_date > today:
                    cell.fill = red_fill  # Ø§Ù„ØªØ§Ø±ÙŠØ® Ø¨Ø§Ù„Ù…Ø³ØªÙ‚Ø¨Ù„
            except:
                pass

if __name__ == '__main__':
    app.run(debug=True)
