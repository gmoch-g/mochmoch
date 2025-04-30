from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from openpyxl.cell.cell import MergedCell
import os

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(BASE_DIR, 'data.xlsx')
@app.route('/')
def index():
    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheets_data = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value if cell.value else '' for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            formatted_row = [{'value': str(cell) if cell is not None else ''} for cell in row]
            rows.append(formatted_row)
        sheets_data.append({'name': sheet, 'headers': headers, 'rows': rows})

    return render_template('index.html', sheets_data=sheets_data)


@app.route('/save', methods=['POST'])
def save():
    try:

        data = request.get_json()

        wb = load_workbook(EXCEL_FILE_PATH)

        for sheet_name, rows in data.items():

            if sheet_name not in wb.sheetnames:
                continue

            sheet = wb[sheet_name]

            max_row = sheet.max_row

            max_column = sheet.max_column

            # امسح المحتوى السابق من الشيت (عدا العناوين في الصف الأول)

            for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):

                for cell in row:

                    if not isinstance(cell, MergedCell):
                        cell.value = None

            # اكتب البيانات الجديدة كما هي (حتى لو الصف فارغ)

            for row_index, row in enumerate(rows[1:], start=2):  # تجاهل العناوين

                for col_index, value in enumerate(row, start=1):

                    cell = sheet.cell(row=row_index, column=col_index)

                    if not isinstance(cell, MergedCell):

                        if value in ("None", None):
                            value = ""

                        cell.value = value

        wb.save(EXCEL_FILE_PATH)

        wb.save(EXCEL_FILE_PATH)

        return jsonify({'status': 'success'})


    except Exception as e:

        print(f"Error while saving: {e}")

        return jsonify({'status': 'error', 'message': str(e)}), 500
    except Exception as e:
        print(f"Error while saving: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    # ألوان الخلايا
    COLORS = {
        'fireRed': 'B71C1C',  # أحمر دموي ناري
        'darkGreen': 'C6EFCE',  # أخضر فاتح
        'white': 'FFFFFF',  # أبيض
        'orange': 'FFEB9C',  # برتقالي فاتح
        'lightGray': 'D9D9D9',  # رمادي فاتح
    }

    # الأعمدة التي فيها تواريخ متابعة أوامر غيار 2025
    allowedColumnsFollowUp = [7, 9, 11, 13, 16, 18, 21]

    for sheet in wb.worksheets:
        sheet_name = sheet.title.strip()

        for row in sheet.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                value = str(cell.value).strip() if cell.value else ''
                valueClean = value.replace("-", "/").replace(" ", "")

                # متابعة مشاريع قيد التنفيذ
                if sheet_name == 'متابعة مشاريع قيد التنفيذ':
                    if idx >= 9:
                        try:
                            val = float(value)
                            if val < 0:
                                cell.fill = PatternFill(start_color=COLORS['fireRed'], end_color=COLORS['fireRed'],
                                                        fill_type='solid')
                            elif val > 0:
                                cell.fill = PatternFill(start_color=COLORS['darkGreen'], end_color=COLORS['darkGreen'],
                                                        fill_type='solid')
                            else:
                                cell.fill = PatternFill(start_color=COLORS['white'], end_color=COLORS['white'],
                                                        fill_type='solid')
                        except ValueError:
                            pass

                # متابعة المشاريع وزارة التخطيط
                elif sheet_name == 'متابعة المشاريع وزارة التخطيط':
                    if idx > 8:
                        prevCell = row[idx - 1] if idx > 0 else None
                        nextCell = row[idx + 1] if idx + 1 < len(row) else None
                        nextNextCell = row[idx + 2] if idx + 2 < len(row) else None

                        if value == "إعلان":
                            cell.fill = PatternFill(start_color=COLORS['darkGreen'], end_color=COLORS['darkGreen'],
                                                    fill_type='solid')
                            if nextCell: nextCell.fill = PatternFill(start_color=COLORS['white'],
                                                                     end_color=COLORS['white'], fill_type='solid')
                            if nextNextCell: nextNextCell.fill = PatternFill(start_color=COLORS['orange'],
                                                                             end_color=COLORS['orange'],
                                                                             fill_type='solid')

                        elif value == "تم":
                            cell.fill = PatternFill(start_color=COLORS['darkGreen'], end_color=COLORS['darkGreen'],
                                                    fill_type='solid')
                            if nextCell: nextCell.fill = PatternFill(start_color=COLORS['orange'],
                                                                     end_color=COLORS['orange'], fill_type='solid')
                            if nextNextCell: nextNextCell.fill = PatternFill(start_color=COLORS['lightGray'],
                                                                             end_color=COLORS['lightGray'],
                                                                             fill_type='solid')

                        elif value == "":
                            if prevCell and str(prevCell.value).strip() == "تم":
                                cell.fill = PatternFill(start_color=COLORS['orange'], end_color=COLORS['orange'],
                                                        fill_type='solid')
                                if nextCell: nextCell.fill = PatternFill(start_color=COLORS['lightGray'],
                                                                         end_color=COLORS['lightGray'],
                                                                         fill_type='solid')
                            else:
                                cell.fill = PatternFill(start_color=COLORS['lightGray'], end_color=COLORS['lightGray'],
                                                        fill_type='solid')

                        else:
                            parts = valueClean.split("/")
                            if len(parts) == 3:
                                try:
                                    cellDate = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                                    if cellDate <= today:
                                        cell.fill = PatternFill(start_color=COLORS['darkGreen'],
                                                                end_color=COLORS['darkGreen'], fill_type='solid')
                                        if nextCell: nextCell.fill = PatternFill(start_color=COLORS['orange'],
                                                                                 end_color=COLORS['orange'],
                                                                                 fill_type='solid')
                                        if nextNextCell: nextNextCell.fill = PatternFill(
                                            start_color=COLORS['lightGray'], end_color=COLORS['lightGray'],
                                            fill_type='solid')
                                    else:
                                        cell.fill = PatternFill(start_color=COLORS['fireRed'],
                                                                end_color=COLORS['fireRed'], fill_type='solid')
                                        if nextCell: nextCell.fill = PatternFill(start_color=COLORS['lightGray'],
                                                                                 end_color=COLORS['lightGray'],
                                                                                 fill_type='solid')
                                except:
                                    pass

                # متابعة أوامر غيار 2025
                elif 'متابعة اوامر غيار 2025' in sheet_name:
                    if idx in allowedColumnsFollowUp:
                        prevCell = row[idx - 1] if idx > 0 else None
                        if prevCell:
                            enteredParts = valueClean.split("/")
                            prevValue = str(prevCell.value).strip().replace("-", "/") if prevCell.value else ''
                            prevParts = prevValue.split("/")

                            try:
                                enteredDate = datetime(int(enteredParts[0]), int(enteredParts[1]), int(enteredParts[2]))
                                prevDate = datetime(int(prevParts[0]), int(prevParts[1]), int(prevParts[2]))
                                if enteredDate > prevDate:
                                    cell.fill = PatternFill(start_color=COLORS['orange'], end_color=COLORS['orange'],
                                                            fill_type='solid')
                                elif enteredDate < prevDate:
                                    cell.fill = PatternFill(start_color=COLORS['fireRed'], end_color=COLORS['fireRed'],
                                                            fill_type='solid')
                                else:
                                    cell.fill = PatternFill(start_color=COLORS['darkGreen'],
                                                            end_color=COLORS['darkGreen'], fill_type='solid')
                            except:
                                pass

if __name__ == '__main__':
    app.run(debug=True)